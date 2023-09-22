import pandas as pd
import openpyxl
import xlwings as xw 

# Well this is trying to figure out how github clones and creates repository's so i dont' 
# Delete the wrong shit. Already at version 1.2. And i am modifiing the changes but not seeing it
# Happen in github desktop. Kind of getting screwed on this. I might need to abandon this set
# it to private and start a new folder becuase It is getting confused
# This is a copy of the first Soca in Tweaking because it goes past 5 channels
# and this is showing up as Modified but why am i not seeing it on Github. 

# Ok looks like i have a starting spot to play with as of 9.22.23




# TODO WORKING TWEAKED THAT CAN HANDLE MORE THAN 5 ENTERED CHANNELS ON THE TOP ONLY SOCA
# TODO BUT NOT MORE THAN 10 BUT ITS A MESS AND REALIZING MY CODING BELOW IT ISN'T QUITE 
# TODO THE SAME AS THE FIRST SOCAPEX(THIS ONE). BRAIN IS CRAMPING ON HOW TO CREATE OOP HAPPEN
# TODO THIS VERSION IS A PYCHARM VERSION. 

path = "/Users/speedy/Documents/GitProjects/AmpCalculate/AmpExcelPython/RigBalanceWorkSheet48way.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['Racks']
wk = xw.books.open(r"RigBalanceWorkSheet48way.xlsx")  # might not work in vsCode not sure what the r is
sheet2 = wk.sheets('Racks')
# speedy = "speedyG"
# sheet2.range("A16").value = speedy # it allows variables which is good for later

# shows what is at B3 on the sheet Racks print(sheet['B3'].value)
# sheet["b3"] = "fucked"  <-- will change that cell, but you have to save it to take
# wb.save(r'path') < have to play with this to see if it works. Try it out on a blank workbook
# df = pd.read_excel(path, engine="openpyxl") is a way to open and use pyxl to edit
# import xlwings as xw check this out it updates a cell without saving
# wk = xw.books.open(r'path')
# sheet = wk.sheets('Racks')
# sheet2.range("A16").value = "hello xlwings" will change the value of the cell immediately.



df = pd.read_excel(path, 'Racks')
wattSheet = pd.read_excel(path, 'Watts').to_dict()
watts = int
actualWatts = []
RowWatts = 0  # to factor that circuit rows watts
volts = wattSheet['Voltage'][0]
lineAmps = []
circuitAmps = []
a = int
cCounter = 1
for counter in range(9):
    cCounter += 1
    channelNums = []
    # this is here to get all 5 of the channel numbers in that circuit
    for k, v in df.items():  # need the items here
        # print(k, v)
        # print(df[k][2])
        # print(f"cCounter {cCounter}")
        channelNums.append(df[k][cCounter])  # this is to create the list of channel numbers
    # print(f"channel numbers: {channelNums}")
    firstSection = channelNums[5:21]  # to slice down to just the channels in the list
    # print(f"firstSection channelNums {firstSection}")
    # print(df[4][2])
    # print(df[8][2])
    # for i in range(len(firstRow)):
    #     print(firstRow[i])
    # looks like I am working to this point, next is to calculate where it falls in the channel scale
    # then to the watts deciding the amps of each channel.
    # use range uses less memory range(100, 100, 1) will use less than getting each number and comparing
    # just have to figure out how to check for the channel number
    # was trying to use the sh1 to create the range lengths wasn't quite getting it correct so the ranges are hard value
    # sh1 = wb['Racks']
    sh2 = wb['Watts']
    rowCount = sh2.max_row
    # print(f"rowcount is: {rowCount}")
    # print(f"Rows watts sheet: {rowCount}")
    # colCount = sh2.max_column
    # print(f"rows: {rowCount}, cols {colCount}.")
    # print(f"rowcount is : {rowCount}")

    try:
        for x in range(16):
            a = firstSection[x]
            # print(f"a variable =: {a}")
            for i in range(rowCount - 2):  # TODO figured out the correct amount for the sh2
                chanLow = wattSheet['ChanLow'][i]
                chanHi = wattSheet['ChanHigh'][i]
                # print(f"low start # {chanLow} high is: {chanHi}")
                chanLow = int(chanLow)
                chanHi = int(chanHi)
                if a in range(chanLow, chanHi):
                    # print("yes it is in the 100 range")
                    watts = wattSheet['Wattage'][i]
                    RowWatts += watts
                    print(f"firstSection channel {a} channelNums {firstSection}")
                    print(f"added watts {watts}, for chanel {a}, total RowWatts: {RowWatts} ")
    except IndexError:
        pass

    lineAmps.append(round((RowWatts / volts / 2), 2))
    actualWatts.append(RowWatts)
    print(f"Total Watts together {actualWatts}")

    # print(f"actual watts this circuit {actualWatts[-1]}")
    # print(f"total watts: {RowWatts} / voltage: {volts} / 2  = {lineAmps[-1]}")
    # # print(lineAmps)
    # print(f"actual watts for each circuit is: {actualWatts}")
    RowWatts = 0  # resets RowWatts for the next circuit check
    # actualWatts = []
# TODO these is a hard value input to the legs have to figure out how to make it relative later

sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("D5", "E5").value = lineAmps[1]
sheet2.range("C6").value = lineAmps[2]
sheet2.range("E6").value = lineAmps[2]
sheet2.range("C7", "D7").value = lineAmps[3]
sheet2.range("D8", "E8").value = lineAmps[4]
sheet2.range("C9").value = lineAmps[5]
sheet2.range("E9").value = lineAmps[5]