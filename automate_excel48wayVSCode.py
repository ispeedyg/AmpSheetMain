import pandas as pd 
import openpyxl
import xlwings as xw

# TODO **** WORKING 48 WAY RACK DONT MESS WITH MAKE A COPY ******
# TODO **** Cannot get the xlwings to work correctly here in VSCODE. Pissing me off.add()


path = "/Users/speedy/Documents/GitProjects/AmpCalculate/AmpExcelPython//RigBalanceWorkSheet48way.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['Racks']
wk = xw.books.open(path) # might not work in vsCode not sure what the r is
sheet2 = wk.sheets('Racks')
# speedy = "speedyG"
# sheet2.range("A16").value = speedy # it allows variables which is good for later

# shows what is at B3 on the sheet Racks print(sheet['B3'].value)
#sheet["b3"] = "DammitJim"  <-- will change that cell, but you have to save it to take
# wb.save(r'path') < have to play with this to see if it works. Try it out on a blank workbook
# df = pd.read_excel(path, engine="openpyxl") is a way to open and use pyxl to edit
#import xlwings as xw check this out it updates a cell without saving
#wk = xw.books.open(r'path')
#sheet = wk.sheets('Racks')
#sheet2.range("A16").value = "hello xlwings" will change the value of the cell immediately.



df = pd.read_excel(path, 'Racks')
wattSheet = pd.read_excel(path, 'Watts').to_dict()
watts = int
actualWatts = []
RowWatts = 0 # to factor that circuit rows watts
volts = wattSheet['Voltage'][0]
lineAmps = []
circuitAmps = []
a = int
cCounter = 1
for counter in range(6):
    cCounter += 1
    channelNums = []
    # this is here to get all 5 of the channel numbers in that circuit
    for k, v in df.items(): # need the items here
        # print(k, v)
        # print(df[k][2])
        # print(f"cCounter {cCounter}")
        channelNums.append(df[k][cCounter]) # this is to create the list of channel numbers
    # print(channelNums)
    firstSection = channelNums[5:10] # to slice down to just the channels in the list

    # print(df[4][2])
    # print(df[8][2])
    # for i in range(len(firstRow)):
    #     print(firstRow[i])
    # looks like I am working to this point, next is to calculate where it falls in the channel scale
    # then to the watts deciding the amps of each channel.
    # use range uses less memory range(100, 100, 1) will use less than getting each number and comparing
    # just have to figure out how to check for the channel number
    # was trying to use the sh1 to create the range lengths wasn't quite getting it correct so the ranges are hard value
    # sh1 = wb['Racks'] 
    sh2 = wb['Watts']
    rowCount = sh2.max_row
    # print(f"Rows watts sheet: {rowCount}")
    # colCount = sh2.max_column
    # print(f"rows: {rowCount}, cols {colCount}.")
    # print(f"rowcount is : {rowCount}")
    for x in range(5):
        a = firstSection[x]
        for i in range(rowCount - 2): # TODO figured out the correct amount for the sh2
            chanLow = wattSheet['ChanLow'][i]
            chanHi = wattSheet['ChanHigh'][i]
            # print(f"low start # {chanLow} high is: {chanHi}")

            if a in range(chanLow, chanHi):
                # print("yes it is in the 100 range")
                watts = wattSheet['Wattage'][i]
                RowWatts += watts
    lineAmps.append(round((RowWatts / volts / 2), 2))
    actualWatts.append(RowWatts)
    # print(firstSection)
    # print(f"actual watts this circuit {actualWatts[-1]}")
    # print(f"total watts: {RowWatts} / voltage: {volts} / 2  = {lineAmps[-1]}")
    # print(lineAmps)
    # print(f"actual watts for each circuit is: {actualWatts}")
    RowWatts = 0 # resets RowWatts for the next circuit check
    # actualWatts = []
# TODO these is a hard value input to the legs have to figure out how to make it relative later

sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("D5", "E5").value = lineAmps[1]
sheet2.range("C6").value = lineAmps[2]
sheet2.range("E6").value = lineAmps[2]
sheet2.range("C7", "D7").value = lineAmps[3]
sheet2.range("D8", "E8").value = lineAmps[4]
sheet2.range("C9").value = lineAmps[5]
sheet2.range("E9").value = lineAmps[5]
# TODO This works for the top half now to create the bottom half. Then next would be to create relative
counter2 = 0
channelNums2 = []
secRowWatts = 0
secLineAmps = []
secVolts = wattSheet['Voltage'][1]
secActualWatts = []
secSection = []
sh2 = wb['Watts']
rowCount = sh2.max_row

for x in range(15, 21, 1):
    for i in range(4, 9, 1):
        channelNums2.append(df.loc[x, i])
        secSection = channelNums2[:5]
        if i == 8:
            movingOn = True
    for c in range(5):
        d = secSection[c]
        for e in range(rowCount - 2):
            secChanLow = wattSheet['ChanLow'][e]
            secChanHi = wattSheet['ChanHigh'][e]
            if d in range(secChanLow, secChanHi):
                secWatts = wattSheet['Wattage'][e]
                secRowWatts += secWatts
    secLineAmps.append(round((secRowWatts / secVolts / 2), 2)) # gets my amps per leg
    secActualWatts.append(secRowWatts)
    # print(secSection)
    # try:
    #     print(f"Actual watts this circuit is {secActualWatts[-1]}") # this circuit total watts
    # except IndexError as errorMessage:
    #     print(f"That key error {errorMessage} Just Clear contents on your own ")
    # else:
    #     pass
    # print(f"Total Watts: {secRowWatts} / Voltage: {secVolts} / 2 Legs = {secLineAmps[-1]}")
    # print(secLineAmps)
    secRowWatts = 0 # resets secRowWatts for the next circuit check
    channelNums2 = []
# TODO -------------------------------- Write to Excel second section ------------------------------------ #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C17", "D17").value = secLineAmps[0]
sheet2.range("D18", "E18").value = secLineAmps[1]
sheet2.range("C19").value = secLineAmps[2]
sheet2.range("E19").value = secLineAmps[2]
sheet2.range("C20", "D20").value = secLineAmps[3]
sheet2.range("D21", "E21").value = secLineAmps[4]
sheet2.range("C22").value = secLineAmps[5]
sheet2.range("E22").value = secLineAmps[5]

# Todo Notes you need to clear contents and save Excel before each section. Have to put in some error catches when there
# is nothing in the voltage to check get a float division by zero error and get an error in the second section when the
# content is cleared get error list index out of range. will have to review on solving that.  But hell just clear the
# contents and build from there.

# TODO Start of the 3rd soca on 24 way Solved the index error above still have other errors will fix in this run

counter3 = 0
channelNums3 = []
thirdRowWatts = 0
thirdLineAmps = []
thirdVolts = wattSheet['Voltage'][2]
thirdActualWatts = []
thirdSection = []

for h in range(27, 33, 1):
    for j in range(4, 9, 1):
        channelNums3.append(df.loc[h, j])
        thirdSection = channelNums3[:5]

    for e in range(5):
        f = thirdSection[e]
        for g in range(rowCount - 2):
            thirdChanLow = wattSheet['ChanLow'][g]
            thirdChanHi = wattSheet['ChanHigh'][g]
            if f in range(thirdChanLow, thirdChanHi):
                thirdWatts = wattSheet['Wattage'][g]
                thirdRowWatts += thirdWatts
    thirdLineAmps.append(round((thirdRowWatts/thirdVolts/2), 2)) # gets my amps per leg
    thirdActualWatts.append(thirdRowWatts)
    # print(thirdSection)
    # print(f"Actual watts this circuit is {thirdActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {thirdRowWatts} / Voltage {thirdVolts} / 2 legs = {thirdLineAmps[-1]}")
    # print(thirdLineAmps)
    thirdRowWatts = 0 # resets thirdRowWatts variable for the next circuit check.
    channelNums3 = [] # resets list so the slice works
# TODO Below is the inserting into the excel Document for the 3rd section Backed up to github as 3rd working
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C29", "D29").value = thirdLineAmps[0]
sheet2.range("D30", "E30").value = thirdLineAmps[1]
sheet2.range("C31").value = thirdLineAmps[2]
sheet2.range("E31").value = thirdLineAmps[2]
sheet2.range("C32", "D32").value = thirdLineAmps[3]
sheet2.range("D33", "E33").value = thirdLineAmps[4]
sheet2.range("C34").value = thirdLineAmps[5]
sheet2.range("E34").value = thirdLineAmps[5]

#TODO ----------------------------  Starting the 4th column of the 24 way ---------------------------------------- #

counter4 = 0
channelNums4 = []
fourthRowWatts = 0
fourthLineAmps = []
fourthVolts = wattSheet['Voltage'][3]
fourthActualWatts = []
fourthSection = []

for k in range(40, 46, 1):
    for m in range(4, 9, 1):
        channelNums4.append(df.loc[k, m])
        fourthSection = channelNums4[:5]

    for n in range(5):
        o = fourthSection[n]
        for p in range(rowCount - 2):
            fourthChanLow = wattSheet['ChanLow'][p]
            fourthChanHi = wattSheet['ChanHigh'][p]
            if o in range(fourthChanLow, fourthChanHi):
                fourthWatts = wattSheet['Wattage'][p]
                fourthRowWatts += fourthWatts
    fourthLineAmps.append(round((fourthRowWatts / fourthVolts / 2), 2)) # Gets my amps per leg
    fourthActualWatts.append(fourthRowWatts)

    # print(fourthSection)
    # print(f"Actual watts this circuit is {fourthActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {fourthRowWatts} / Voltage: {fourthVolts} / 2 legs = {fourthLineAmps[-1]}")
    # print(fourthLineAmps)
    fourthRowWatts = 0 # resets variable for the next circuit check
    channelNums4 = [] # resets so the slice works
# TODO --------------------- This works for the 4th section now to insert it into the Excel -------------------- #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C42", "D42").value = fourthLineAmps[0]
sheet2.range("D43", "E43").value = fourthLineAmps[1]
sheet2.range("C44").value = fourthLineAmps[2]
sheet2.range("E44").value = fourthLineAmps[2]
sheet2.range("C45", "D45").value = fourthLineAmps[3]
sheet2.range("D46", "E46").value = fourthLineAmps[4]
sheet2.range("C47").value = fourthLineAmps[5]
sheet2.range("E47").value = fourthLineAmps[5]
# TODO ------ This works for the 24 way rack save as a separate sheet and next create the 48 way sheet --------- #
counter5 = 0
channelNums5 = []
fifthRowWatts = 0
fifthLineAmps = []
fifthVolts = wattSheet['Voltage'][4]
fifthActualWatts = []
fifthSection = []

for aa in range(53, 59, 1):
    for bb in range(4, 9, 1):
        channelNums5.append(df.loc[aa, bb])
        fifthSection = channelNums5[:5]

    for cc in range(5):
        dd = fifthSection[cc]
        for ee in range(rowCount - 2):

            fifthChanLow = wattSheet['ChanLow'][ee]
            fifthChanHi = wattSheet['ChanHigh'][ee]
            if dd in range(fifthChanLow, fifthChanHi):
                fifthWatts = wattSheet['Wattage'][ee]
                # print(f"chan: {dd} watts = {fifthWatts}")
                fifthRowWatts += fifthWatts
    fifthLineAmps.append(round((fifthRowWatts / fifthVolts / 2), 2)) # Gets my amps per leg
    fifthActualWatts.append(fifthRowWatts)

    # print(fifthSection)
    # print(f"Actual watts this circuit is {fifthActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {fifthRowWatts} / Voltage: {fifthVolts} / 2 legs = {fifthLineAmps[-1]}")
    # print(fifthLineAmps)
    fifthRowWatts = 0 # resets variable for the next circuit check
    channelNums5 = [] # resets so the slice works
#TODO ---------------------  THE 5TH SOCA IS COMPLETE BELOW ENTERING THE EXCEL ENTRIES ------------------ #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C55", "D55").value = fifthLineAmps[0]
sheet2.range("D56", "E56").value = fifthLineAmps[1]
sheet2.range("C57").value = fifthLineAmps[2]
sheet2.range("E57").value = fifthLineAmps[2]
sheet2.range("C58", "D58").value = fifthLineAmps[3]
sheet2.range("D59", "E59").value = fifthLineAmps[4]
sheet2.range("C60").value = fifthLineAmps[5]
sheet2.range("E60").value = fifthLineAmps[5]
# TODO --------------------- Below is the 6th soca entry -------------------------------------------- #
counter6 = 0
channelNums6 = []
sixthRowWatts = 0
sixthLineAmps = []
sixthVolts = wattSheet['Voltage'][5]
sixthActualWatts = []
sixthSection = []

for ff in range(66, 72, 1):
    for gg in range(4, 9, 1):
        channelNums6.append(df.loc[ff, gg])
        sixthSection = channelNums6[:5]

    for hh in range(5):
        ii = sixthSection[hh]
        for jj in range(rowCount - 2):

            sixthChanLow = wattSheet['ChanLow'][jj]
            sixthChanHi = wattSheet['ChanHigh'][jj]
            if ii in range(sixthChanLow, sixthChanHi):
                sixthWatts = wattSheet['Wattage'][jj]
                # print(f"chan: {ii} watts = {sixthWatts}")
                sixthRowWatts += sixthWatts
    sixthLineAmps.append(round((sixthRowWatts / sixthVolts / 2), 2)) # Gets my amps per leg
    sixthActualWatts.append(sixthRowWatts)

    # print(sixthSection)
    # print(f"Actual watts this circuit is {sixthActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {sixthRowWatts} / Voltage: {sixthVolts} / 2 legs = {sixthLineAmps[-1]}")
    # print(sixthLineAmps)
    sixthRowWatts = 0 # resets variable for the next circuit check
    channelNums6 = [] # resets so the slice works
#TODO ----------- Below is the Excel Entries for 6th soca ------------------------------------------ #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C68", "D68").value = sixthLineAmps[0]
sheet2.range("D69", "E69").value = sixthLineAmps[1]
sheet2.range("C70").value = sixthLineAmps[2]
sheet2.range("E70").value = sixthLineAmps[2]
sheet2.range("C71", "D71").value = sixthLineAmps[3]
sheet2.range("D72", "E72").value = sixthLineAmps[4]
sheet2.range("C73").value = sixthLineAmps[5]
sheet2.range("E73").value = sixthLineAmps[5]
#TODO --------------------- Below is the 7th soca entry ------------------------------------------ #
counter7 = 0
channelNums7 = []
sevRowWatts = 0
sevLineAmps = []
sevVolts = wattSheet['Voltage'][6]
sevActualWatts = []
sevSection = []

for kk in range(79, 85, 1):
    for ll in range(4, 9, 1):
        channelNums7.append(df.loc[kk, ll])
        sevSection = channelNums7[:5]

    for mm in range(5):
        nn = sevSection[mm]
        for oo in range(rowCount - 2):

            sevChanLow = wattSheet['ChanLow'][oo]
            sevChanHi = wattSheet['ChanHigh'][oo]
            if nn in range(sevChanLow, sevChanHi):
                sevWatts = wattSheet['Wattage'][oo]
                # print(f"chan: {ii} watts = {sixthWatts}")
                sevRowWatts += sevWatts
    sevLineAmps.append(round((sevRowWatts / sevVolts / 2), 2)) # Gets my amps per leg
    sevActualWatts.append(sevRowWatts)

    # print(sevSection)
    # print(f"Actual watts this circuit is {sevActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {sevRowWatts} / Voltage: {sevVolts} / 2 legs = {sevLineAmps[-1]}")
    # print(sevLineAmps)
    sevRowWatts = 0 # resets variable for the next circuit check
    channelNums7 = [] # resets so the slice works
#TODO ---------------------- Below is the Excel Entries for 7th Soca --------------------------- #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C81", "D81").value = sevLineAmps[0]
sheet2.range("D82", "E82").value = sevLineAmps[1]
sheet2.range("C83").value = sevLineAmps[2]
sheet2.range("E83").value = sevLineAmps[2]
sheet2.range("C84", "D84").value = sevLineAmps[3]
sheet2.range("D85", "E85").value = sevLineAmps[4]
sheet2.range("C86").value = sevLineAmps[5]
sheet2.range("E86").value = sevLineAmps[5]
#TODO ---------------------- Below is Soca #8 calculations -------------------------------------#
counter8 = 0
channelNums8 = []
eightRowWatts = 0
eightLineAmps = []
eightVolts = wattSheet['Voltage'][7]
eightActualWatts = []
eightSection = []

for pp in range(92, 98, 1):
    for qq in range(4, 9, 1):
        channelNums8.append(df.loc[pp, qq])
        eightSection = channelNums8[:5]

    for rr in range(5):
        ss = eightSection[rr]
        for tt in range(rowCount - 2):

            eightChanLow = wattSheet['ChanLow'][tt]
            eightChanHi = wattSheet['ChanHigh'][tt]
            if ss in range(eightChanLow, eightChanHi):
                eightWatts = wattSheet['Wattage'][tt]
                # print(f"chan: {ii} watts = {sixthWatts}")
                eightRowWatts += eightWatts
    eightLineAmps.append(round((eightRowWatts / eightVolts / 2), 2)) # Gets my amps per leg
    eightActualWatts.append(eightRowWatts)

    # print(eightSection)
    # print(f"Actual watts this circuit is {eightActualWatts[-1]}") # this circuits total watts
    # print(f"Total Watts: {eightRowWatts} / Voltage: {eightVolts} / 2 legs = {eightLineAmps[-1]}")
    # print(eightLineAmps)
    eightRowWatts = 0 # resets variable for the next circuit check
    channelNums8 = [] # resets so the slice works
#TODO ----------------- Below is the EXCEL entries for soca 8 ------------------------------------------------- #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C94", "D94").value = eightLineAmps[0]
sheet2.range("D95", "E95").value = eightLineAmps[1]
sheet2.range("C96").value = eightLineAmps[2]
sheet2.range("E96").value = eightLineAmps[2]
sheet2.range("C97", "D97").value = eightLineAmps[3]
sheet2.range("D98", "E98").value = eightLineAmps[4]
sheet2.range("C99").value = eightLineAmps[5]
sheet2.range("E99").value = eightLineAmps[5]
#TODO # -------------- 48 channel rack completed. Down the road create OOP for a lot less typing and put in a safety #
#TODO # --- for if a number is typed that is not at all in the watt sheet. Plus protect against letters or symbols.
#TODO # - Which won't happen for a while but this is a great practice for learning python and excel just added it
# a couple weeks ago which I will be able to put this script into where VBA is and keep it as one file 9/2/2023