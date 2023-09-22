import pandas as pd
import openpyxl
import xlwings as xw

# isn't that peachy. 1st run on this cloud thing i get it too work. 2nd run get the xlwings error again
# fuck me
path = "/Users/speedy/Documents/GitProjects/AmpCalculate/AmpExcelPython/RigBalanceWorkSheet.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['Racks']
wk = xw.books.open(path)
# might not work in vsCode not sure what the r is
sheet2 = wk.sheets('Racks')
df = pd.read_excel(path, 'Racks')
wattSheet = pd.read_excel(path, 'Watts').to_dict()
# speedy = "speedyG"
# sheet2.range("A16").value = speedy # it allows variables which is good for later

# shows what is at B3 on the sheet Racks print(sheet['B3'].value)
# sheet["b3"] = "fucked"  <-- will change that cell, but you have to save it to take
# wb.save(r'path') < have to play with this to see if it works. Try it out on a blank workbook
# df = pd.read_excel(path, engine="openpyxl") is a way to open and use pyxl to edit
# import xlwings as xw check this out it updates a cell without saving
# wk = xw.books.open(r'path')
# sheet = wk.sheets('Racks')
# sheet2.range("A16").value = "hello xlwings" will change the value of the cell immediately.
# TODO just keeping these notes above for referencing what it took to build
counter = 0
channelNums = []
RowWatts = 0
lineAmps = []
Volts = wattSheet['Voltage'][0]
ActualWatts = []
Section = []
sh2 = wb['Watts']
rowCount = sh2.max_row

for h in range(2, 8, 1):
    for j in range(4, 9, 1):
        channelNums.append(df.loc[h, j])
        Section = channelNums[:5]

    for e in range(5):
        f = Section[e]
        for g in range(rowCount - 2):
            ChanLow = wattSheet['ChanLow'][g]
            ChanHi = wattSheet['ChanHigh'][g]
            if f in range(ChanLow, ChanHi):
                Watts = wattSheet['Wattage'][g]
                RowWatts += Watts
    lineAmps.append(round((RowWatts/Volts/2), 2))  # gets my amps per leg
    ActualWatts.append(RowWatts)
    print(Section)
    print(f"Actual watts this circuit is {ActualWatts[-1]}")  # this circuits total watts
    print(f"Total Watts: {RowWatts} / Voltage {Volts} / 2 legs = {lineAmps[-1]}")
    print(lineAmps)
    RowWatts = 0  # resets RowWatts variable for the next circuit check.
    channelNums = []  # resets list so the slice works
# TODO these is a hard value input to the legs have to figure out how to make it relative later

sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("D5", "E5").value = lineAmps[1]
sheet2.range("C6").value = lineAmps[2]
sheet2.range("E6").value = lineAmps[2]
sheet2.range("C7", "D7").value = lineAmps[3]
sheet2.range("D8", "E8").value = lineAmps[4]
sheet2.range("C9").value = lineAmps[5]
sheet2.range("E9").value = lineAmps[5]
# TODO This works for the top half now to create the bottom half. Then next would be to create relative
counter2 = 0
channelNums2 = []
secRowWatts = 0
secLineAmps = []
secVolts = wattSheet['Voltage'][1]
secActualWatts = []
secSection = []
sh2 = wb['Watts']
rowCount = sh2.max_row

for x in range(15, 21, 1):
    for i in range(4, 9, 1):
        channelNums2.append(df.loc[x, i])
        secSection = channelNums2[:5]

    for c in range(5):
        d = secSection[c]
        for e in range(rowCount - 1):
            secChanLow = wattSheet['ChanLow'][e]
            secChanHi = wattSheet['ChanHigh'][e]
            if d in range(secChanLow, secChanHi):
                secWatts = wattSheet['Wattage'][e]
                secRowWatts += secWatts
    secLineAmps.append(round((secRowWatts / secVolts / 2), 2))  # gets my amps per leg
    secActualWatts.append(secRowWatts)
    print(secSection)
    try:
        print(f"Actual watts this circuit is {secActualWatts[-1]}")  # this circuit total watts
    except IndexError as errorMessage:
        print(f"That key error {errorMessage} Just Clear contents on your own ")
    else:
        pass
    print(f"Total Watts: {secRowWatts} / Voltage: {secVolts} / 2 Legs = {secLineAmps[-1]}")
    print(secLineAmps)
    secRowWatts = 0  # resets secRowWatts for the next circuit check
    channelNums2 = []
# TODO -------------------------------- Write to Excel second section ------------------------------------ #
# example sheet2.range("C4", "D4").value = lineAmps[0]
sheet2.range("C17", "D17").value = secLineAmps[0]
sheet2.range("D18", "E18").value = secLineAmps[1]
sheet2.range("C19").value = secLineAmps[2]
sheet2.range("E19").value = secLineAmps[2]
sheet2.range("C20", "D20").value = secLineAmps[3]
sheet2.range("D21", "E21").value = secLineAmps[4]
sheet2.range("C22").value = secLineAmps[5]
sheet2.range("E22").value = secLineAmps[5]

# Todo Notes you need to clear contents and save Excel before each section. Have to put in some error catches when there
# is nothing in the voltage to check get a float division by zero error and get an error in the second section when the
# content is cleared get error list index out of range. will have to review on solving that.  But hell just clear the
# contents and build from there.
